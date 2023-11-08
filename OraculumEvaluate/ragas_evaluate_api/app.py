## app.py

from flask import Flask, request, jsonify, send_from_directory
from data_processor import DataProcessor
from models import db, Output
from config import Config
import json, os
from openpyxl import Workbook
        
class FlaskApp:
    """FlaskApp class for setting up and running the Flask app."""
    def __init__(self, config: Config):
        self.app = Flask(__name__)
        self.app.config[config.FLASK_APP_SETTINGS] = config.SECRET_KEY
        self.app.config['SQLALCHEMY_DATABASE_URI'] = config.SQLALCHEMY_DATABASE_URI
        self.app.config['SECURITY_PASSWORD_SALT'] = config.SECURITY_PASSWORD_SALT
        db.init_app(self.app)
        self.data_processor = DataProcessor()

    def run(self):
        """Run the Flask app."""

        @self.app.route('/augment', methods=['POST'])
        def augment():
            """Augment JSON data."""
            input_data = request.get_json()
            augmented_data = self.data_processor.augment_data(input_data)
            output = Output(json.dumps(augmented_data), "augmented")
            db.session.add(output)
            db.session.commit()
            return jsonify(output.to_dict()), 200
        
        @self.app.route('/evaluate', methods=['GET'])
        def evaluate():
            """Evaluate data by id."""
            id = int(request.args.get('id'))
            data = Output.query.get({"id": id})
            # json_data = json.loads(data)
            evaluated_data = self.data_processor.eval_data(json.loads(data.output_data))
            output = Output(json.dumps(evaluated_data), "evaluated")
            db.session.add(output)
            db.session.commit()
            return jsonify(output.to_dict()), 200


        @self.app.route('/retrieve', methods=['GET'])
        def retrieve():
            """Retrieve processed data."""
            id_str = request.args.get('id')
            if id_str:
                if not id_str.isdigit():
                    return "Invalid ID provided", 400
                id = int(id_str)
                data = Output.query.get(id)
                if data is None:
                    return "Record not found", 404
                return jsonify(data.to_dict()), 200
            else:
                outputs = Output.query.all()
                return jsonify([output.to_dict() for output in outputs]), 200

        @self.app.route('/report', methods = ['GET'])
        def report():
            """Save report as xlsx by id."""
            id_str = request.args.get('id')
            if not id_str or not id_str.isdigit():
                print("Invalid ID provided")  # Debugging statement
                return "Invalid ID provided", 400

            id = int(id_str)
            data = Output.query.get(id)
            
            if data is None:
                print(f"Record with ID {id} not found")  # Debugging statement
                return "Record not found", 404

                # Parsing the output_data as JSON
            try:
                parsed_data = json.loads(data.output_data)
                flat_json = []
                total_metrics = None
                for item in parsed_data:
                    if "question" in item and "answer" in item and "contexts" in item and "evaluation" in item:
                        faithfulness = round(item["evaluation"].get("faithfulness", 0), 2)
                        answer_relevancy = round(item["evaluation"].get("answer_relevancy", 0), 2)
                        context_precision = round(item["evaluation"].get("context_precision", 0), 2)
                        context_recall = round(item["evaluation"].get("context_recall", 0), 2)
                        ragas_score = round(item["evaluation"].get("ragas_score", 0), 2)
                        ground_truths = item.get("ground_truths", "")
                        rounded_item = {
                            "question": item["question"],
                            "answer": item["answer"],
                            "contexts": ' *** '.join(item["contexts"]),
                            "ground_truths": ground_truths,
                            "faithfulness": faithfulness,
                            "answer_relevancy": answer_relevancy,
                            "context_precision": context_precision,
                            "context_recall": context_recall,
                            "ragas_score": ragas_score
                        }
                        flat_json.append(rounded_item)
                else:
                    total_metrics = item
                if total_metrics:
                    flat_json.append(total_metrics)
                # with open('log.json', 'w', encoding='utf-8') as f:
                #     json.dump(flat_json, f)
            except json.JSONDecodeError:
                print("Error decoding the JSON data")
                return "Error processing data", 500

            # Initialize a new Excel workbook and worksheet
            wb = Workbook()
            ws = wb.active


            # Write headers (assuming the first item's keys are representative)
            headers = flat_json[0].keys()
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)

            # Write data
            for row_num, item in enumerate(flat_json, 2):  # Start from the second row as the first row is headers
                for col_num, header in enumerate(headers, 1):
                    value = item.get(header, "")  # Use the get method with a default value of empty string
                    if isinstance(value, list):
                        value = ", ".join(value)
                    elif isinstance(value, dict):
                        value = ", ".join(f"{k}: {v}" for k, v in value.items())
                    ws.cell(row=row_num, column=col_num, value=value)

            filename = f"report_{id}.xlsx"
            filepath = os.path.abspath(f"./{filename}")  # Convert to absolute path

            wb.save(filepath)
            
            try:
                return send_from_directory(directory=os.path.dirname(filepath), path=os.path.basename(filepath), as_attachment=True)
            except Exception as e:
                print(f"Error sending the file: {e}")  # Debugging statement
                return "Error sending the file", 500

            
            

        self.app.run()

if __name__ == "__main__":
    config = Config()
    flask_app = FlaskApp(config)
    flask_app.run()
